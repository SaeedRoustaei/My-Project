from openpyxl import load_workbook

EX="/home/agha-saeed/Downloads/Port Plan - MCI-PMR-20240123-00001.xlsx"
workbook = load_workbook(EX, read_only=True)
sheet = workbook["Sheet13"]

# final="config firewall address" + "\n"
# for row in sheet.iter_rows(values_only=True):
#       ip = row [0].strip()
#       final=final+ "edit " + ip + "/32" + "\n" + "set subent " + ip + "/32" + "\n" +"next" + "\n"
# print (final) 


# final="config firewall addrgrp" + "\n" + "edit CRM-Nodes" + "\n" + "set member "
# for row in sheet.iter_rows(values_only=True):
#       ip = row [0].strip()
#       final=final+ ip + "/32 "
# print (final)
      

for row in sheet.iter_rows(values_only=True):
      interface = row [0]
      Rack = row [1]
    #  Server = row [2].strip()
    #  server_name = row [2]
      server_port = row [2]
      Role= row [3]
      command = f"interface eth 1/{interface}" + "\n" + "description To-Churn-" + f'{Rack}-{server_port}-{Role}'
      print (command)


